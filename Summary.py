from datetime import date
from openpyxl.utils import get_column_letter

from Constants import (LOG_SHEET, SUM_SHEET, TARGET_HOURS,
                       C_HEADER, C_HEADER_SUM, C_OVER, C_UNDER, C_TODAY, C_WHITE)
from Styles import style_cell
from Workbook import read_log


def build_summary(wb):
    """Rebuild the Summary sheet from scratch based on the current log."""
    if SUM_SHEET in wb.sheetnames:
        del wb[SUM_SHEET]
    ws = wb.create_sheet(SUM_SHEET)
    ws_log = wb[LOG_SHEET]

    days = read_log(ws_log)
    if not days:
        ws["A1"] = "No completed sessions yet."
        return

    _write_daily_section(ws, days)
    _write_weekly_pivot(ws, days, start_row=len(days) + 4)
    ws.freeze_panes = "A2"


def _row_color(is_over):
    return C_OVER if is_over else C_UNDER


def _decimal_to_hhmm(hours):
    """Convert decimal hours to HH:MM format."""
    total_minutes = round(hours * 60)
    h = total_minutes // 60
    m = total_minutes % 60
    return f"{h}:{m:02d}"


def _write_daily_section(ws, days):
    headers = ["Date", "Week", "Sessions", "Total Hours (100th)", "Total Hours (60th)", "vs 7.40h", "Status"]
    col_widths = [14, 8, 12, 14, 16, 14, 14]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        style_cell(cell, bold=True, color=C_HEADER, font_color=C_WHITE, size=11)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22

    for i, (date_str, info) in enumerate(sorted(days.items()), start=2):
        total = info["total"]
        diff = total - TARGET_HOURS
        is_over = diff >= 0
        sign = "+" if is_over else ""

        values = [
            date_str, info["week"], len(info["sessions"]),
            round(total, 2), _decimal_to_hhmm(total), f"{sign}{diff:.2f}h",
            "✅ On track" if is_over else "⚠️  Under",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=col, value=val)
            style_cell(cell, color=_row_color(is_over),
                       num_fmt="0.00" if col == 4 else None)


def _write_weekly_pivot(ws, days, start_row):
    # Section label
    label = ws.cell(row=start_row - 1, column=1, value="Weekly Summary")
    style_cell(label, bold=True, color=C_HEADER_SUM, font_color=C_WHITE, size=11)
    ws.merge_cells(start_row=start_row - 1, start_column=1,
                   end_row=start_row - 1, end_column=6)

    headers = ["Week", "Days Worked", "Total Hours", "Target Hours", "Delta", "Status"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=h)
        style_cell(cell, bold=True, color=C_HEADER_SUM, font_color=C_WHITE, size=11)
    ws.row_dimensions[start_row].height = 22

    # Aggregate by week
    weeks = {}
    for date_str, info in days.items():
        w = info["week"]
        if w not in weeks:
            weeks[w] = {"days": 0, "total": 0.0}
        weeks[w]["days"] += 1
        weeks[w]["total"] += info["total"]

    for j, (week_num, winfo) in enumerate(sorted(weeks.items()), start=1):
        total = winfo["total"]
        target = winfo["days"] * TARGET_HOURS
        diff = total - target
        is_over = diff >= 0
        sign = "+" if is_over else ""
        color = C_OVER if is_over else C_UNDER

        values = [
            week_num, winfo["days"], round(total, 2), round(target, 2),
            f"{sign}{diff:.2f}h", "✅ On track" if is_over else "⚠️  Under",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=start_row + j, column=col, value=val)
            style_cell(cell, color=color, num_fmt="0.00" if col in (3, 4) else None)
