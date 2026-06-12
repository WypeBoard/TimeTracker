from datetime import datetime, date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import FORMULAE

from Constants import EXCEL_FILE, LOG_SHEET, SUM_SHEET, C_HEADER, C_ALT, C_WHITE
from Styles import thin_border, style_cell


def create_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = LOG_SHEET

    for col, h in enumerate(["Date", "Week", "Start", "End", "Hours", "Task/Epic"], 1):
        cell = ws.cell(row=1, column=col, value=h)
        style_cell(cell, bold=True, color=C_HEADER, font_color=C_WHITE, size=11)

    ws.row_dimensions[1].height = 22
    for col, w in zip("ABCDEF", [14, 8, 12, 12, 10, 18]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    wb.create_sheet(SUM_SHEET)
    wb.save(EXCEL_FILE)
    print(f"Created {EXCEL_FILE}")
    return wb


def ensure_task_column(ws):
    """Add Task/Epic header in column F if it is missing (for pre-existing workbooks)."""
    header = ws.cell(row=1, column=6).value
    if header != "Task/Epic":
        cell = ws.cell(row=1, column=6, value="Task/Epic")
        style_cell(cell, bold=True, color=C_HEADER, font_color=C_WHITE, size=11)
        ws.column_dimensions["F"].width = 18


def format_log_row(ws, row):
    """Apply styling and number formats. Week formula is written here last
    to prevent openpyxl from prefixing it with @ when the cell has a text format."""
    fill = PatternFill("solid", start_color=C_ALT) if row % 2 == 0 else None
    for col in range(1, 7):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(name="Arial", size=10)
        cell.border = thin_border()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if fill:
            cell.fill = fill

    ws.cell(row=row, column=5).number_format = "0.00"
    # Force Start/End to plain text so Excel never converts them to time serials
    ws.cell(row=row, column=3).number_format = "@"
    ws.cell(row=row, column=4).number_format = "@"


def find_open_session(ws):
    """Return the row index of the most recent unclosed session, or None."""
    for row in range(ws.max_row, 1, -1):
        if ws.cell(row=row, column=3).value and not ws.cell(row=row, column=4).value:
            return row
    return None


def find_today_rows(ws):
    """Return a list of row indices (in order) for today's sessions."""
    today_str = date.today().strftime("%Y-%m-%d")
    rows = []
    for row in range(2, ws.max_row + 1):
        date_val = ws.cell(row=row, column=1).value
        if date_val and str(date_val).strip() == today_str:
            rows.append(row)
    return rows


def _to_hhmm(val):
    """Normalise a cell value to 'HH:MM' string regardless of type."""
    if val is None:
        return None
    if hasattr(val, "hour"):  # datetime.time or datetime.datetime
        return f"{val.hour:02d}:{val.minute:02d}"
    s = str(val).strip()
    try:
        f = float(s)
        total_minutes = round(f * 24 * 60)
        return f"{total_minutes // 60:02d}:{total_minutes % 60:02d}"
    except ValueError:
        return s  # already a plain "HH:MM" string


def read_log(ws):
    """
    Parse the log sheet into a per-day structure.
    Returns { date_str: {"week": int, "sessions": [(start, end, task)], "total": float} }
    Only rows with both Start and End are included.
    """
    days = {}
    for row in range(2, ws.max_row + 1):
        date_val = ws.cell(row=row, column=1).value
        week_val = datetime.strptime(date_val, "%Y-%m-%d").isocalendar().week
        start_val = _to_hhmm(ws.cell(row=row, column=3).value)
        end_val   = _to_hhmm(ws.cell(row=row, column=4).value)
        task_val  = ws.cell(row=row, column=6).value or ""

        if not (date_val and start_val and end_val):
            continue

        date_str = str(date_val).strip()
        try:
            s = datetime.strptime(start_val, "%H:%M")
            e = datetime.strptime(end_val, "%H:%M")
            hours = (e.hour * 60 + e.minute - (s.hour * 60 + s.minute)) / 60
        except ValueError:
            hours = 0.0

        if date_str not in days:
            days[date_str] = {"week": week_val, "sessions": [], "total": 0.0}
        days[date_str]["sessions"].append((start_val, end_val, str(task_val).strip()))
        days[date_str]["total"] += hours

    return days