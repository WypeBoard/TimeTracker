import os
from datetime import datetime, date
from openpyxl import load_workbook

from Constants import EXCEL_FILE, LOG_SHEET, TARGET_HOURS
from Workbook import (
    create_workbook, ensure_task_column, format_log_row,
    find_open_session, find_today_rows, read_log,
)
from Summary import build_summary
from Promark import build_promark, promark_entry
from Printer import DayStatus, print_status, print_day_summary, print_promark_week, print_log_week

_NO_FILE_MSG = "No timetracker file found. Run 'start' first."


def cmd_start(time_override=None, epic=None):
    if not os.path.exists(EXCEL_FILE):
        create_workbook()

    wb = load_workbook(EXCEL_FILE)
    ws = wb[LOG_SHEET]
    ensure_task_column(ws)

    open_row = find_open_session(ws)
    if open_row:
        t = ws.cell(row=open_row, column=3).value
        print(f"⚠  Session already running since {t}. Use 'pause' or 'stop' first.")
        return

    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    time_str = time_override or now.strftime("%H:%M")
    row = ws.max_row + 1

    ws.cell(row=row, column=1, value=date_str)
    ws.cell(row=row, column=2, value=f'=WEEKNUM(A{row})')
    ws.cell(row=row, column=3, value=time_str)
    ws.cell(row=row, column=4, value=None)
    ws.cell(row=row, column=5, value=f'=IF(D{row}<>"",ROUND((D{row}-C{row})*24, 2),"")')
    ws.cell(row=row, column=6, value=epic or "")
    format_log_row(ws, row)

    wb.save(EXCEL_FILE)
    epic_suffix = f" — {epic}" if epic else ""
    note = " (manually set)" if time_override else ""
    print(f"▶  Started at {time_str}{epic_suffix}{note}")


def cmd_restart(time_override=None):
    """Close the current session and open a new one, carrying the epic forward."""
    if not os.path.exists(EXCEL_FILE):
        print(_NO_FILE_MSG)
        return

    wb = load_workbook(EXCEL_FILE)
    ws = wb[LOG_SHEET]
    ensure_task_column(ws)

    open_row = find_open_session(ws)
    if not open_row:
        print("No open session found. Run 'start' first.")
        return

    # Read data from the open session before closing it
    start_val = ws.cell(row=open_row, column=3).value
    task_val = str(ws.cell(row=open_row, column=6).value or "").strip()
    today_rows = find_today_rows(ws)
    session_num = today_rows.index(open_row) + 1 if open_row in today_rows else "?"

    time_str = time_override or datetime.now().strftime("%H:%M")
    ws.cell(row=open_row, column=4, value=time_str)
    wb.save(EXCEL_FILE)
    print(f"⏸  Closed #{session_num} — {start_val} → {time_str}")

    # Open a new session carrying the task forward
    wb = load_workbook(EXCEL_FILE)
    ws = wb[LOG_SHEET]
    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    row = ws.max_row + 1

    ws.cell(row=row, column=1, value=date_str)
    ws.cell(row=row, column=2, value=f'=WEEKNUM(A{row})')
    ws.cell(row=row, column=3, value=time_str)
    ws.cell(row=row, column=4, value=None)
    ws.cell(row=row, column=5, value=f'=IF(D{row}<>"",ROUND((D{row}-C{row})*24, 2),"")')
    ws.cell(row=row, column=6, value=task_val)
    format_log_row(ws, row)
    wb.save(EXCEL_FILE)

    new_num = session_num + 1 if isinstance(session_num, int) else "?"
    carry_note = f"  {task_val}  (carried forward)" if task_val else ""
    print(f"▶  Started #{new_num} — {time_str}{carry_note}")


def cmd_pause(time_override=None):
    _close_session(time_override)


def cmd_stop(time_override=None):
    wb, ws = _close_session(time_override)
    if wb is None:
        return

    print_day_summary(_build_day_status(ws))
    print("🔄  Rebuilding Summary sheet…")
    build_summary(wb)
    print("🔄  Rebuilding Promark sheet…")
    build_promark(wb)
    wb.save(EXCEL_FILE)
    print("✅  Done — Summary and Promark tabs updated.")


def cmd_status():
    if not os.path.exists(EXCEL_FILE):
        print(_NO_FILE_MSG)
        return

    wb = load_workbook(EXCEL_FILE)
    ws = wb[LOG_SHEET]
    now = datetime.now()
    print_status(_build_day_status(ws, now), now)


def cmd_task(epic, session_num=None):
    """Tag a session row with a Task/Epic ID."""
    if not os.path.exists(EXCEL_FILE):
        print(_NO_FILE_MSG)
        return

    wb = load_workbook(EXCEL_FILE)
    ws = wb[LOG_SHEET]
    ensure_task_column(ws)

    today_rows = find_today_rows(ws)
    if not today_rows:
        print("No sessions found for today.")
        return

    if session_num is not None:
        # -s N selects the Nth session (1-indexed)
        idx = session_num - 1
        if idx < 0 or idx >= len(today_rows):
            print(f"⚠  Session #{session_num} not found. Today has {len(today_rows)} session(s).")
            return
        target_row = today_rows[idx]
    else:
        # Default: last open session, or last closed session
        open_row = find_open_session(ws)
        target_row = open_row if open_row and open_row in today_rows else today_rows[-1]

    target_idx = today_rows.index(target_row) + 1
    ws.cell(row=target_row, column=6, value=epic)
    wb.save(EXCEL_FILE)
    print(f"📌  #{target_idx}  {epic}  saved.")


def cmd_log(week_str=None):
    """Show a full week's sessions with task labels (replaces promark terminal output)."""
    if not os.path.exists(EXCEL_FILE):
        print(_NO_FILE_MSG)
        return

    wb = load_workbook(EXCEL_FILE)
    ws = wb[LOG_SHEET]
    days = read_log(ws)

    today_iso = date.today().isocalendar()
    if week_str is None:
        target_week = today_iso.week
        target_year = today_iso.year
    else:
        target_week = int(week_str[1:])
        target_year = today_iso.year

    week_days = {
        d: info for d, info in days.items()
        if date.fromisoformat(d).isocalendar().week == target_week
        and date.fromisoformat(d).isocalendar().year == target_year
    }

    print_log_week(week_days, target_week, target_year)


def cmd_promark(week_str=None):
    """Print a full week's Promark entries to the terminal and rebuild the Promark tab."""
    if not os.path.exists(EXCEL_FILE):
        print(_NO_FILE_MSG)
        return

    wb = load_workbook(EXCEL_FILE)
    ws = wb[LOG_SHEET]
    days = read_log(ws)

    today_iso = date.today().isocalendar()
    if week_str is None:
        target_week = today_iso.week
        target_year = today_iso.year
    else:
        target_week = int(week_str[1:])
        target_year = today_iso.year

    week_days = {
        d: info for d, info in days.items()
        if date.fromisoformat(d).isocalendar().week == target_week
        and date.fromisoformat(d).isocalendar().year == target_year
    }

    print_promark_week(week_days, target_week, target_year)

    print("🔄  Rebuilding Promark sheet…")
    build_promark(wb)
    wb.save(EXCEL_FILE)
    print("✅  Promark tab updated.")


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _build_day_status(ws, now: datetime | None = None) -> DayStatus:
    """Gather today's log data from the worksheet into a DayStatus."""
    now = now or datetime.now()
    today_str = date.today().strftime("%Y-%m-%d")
    days = read_log(ws)
    info = days.get(today_str)

    sessions = info["sessions"] if info else []   # list of (start, end, task)
    total_hours = info["total"] if info else 0.0

    active_start = None
    active_hours = 0.0
    active_task = None
    open_row = find_open_session(ws)
    if open_row:
        raw = ws.cell(row=open_row, column=3).value
        active_start = str(raw).strip() if raw else None
        active_task = str(ws.cell(row=open_row, column=6).value or "").strip() or None
        if active_start:
            try:
                s = datetime.strptime(active_start, "%H:%M")
                active_hours = (now.hour * 60 + now.minute - (s.hour * 60 + s.minute)) / 60
            except ValueError:
                active_hours = 0.0

    pm_start, pm_end = promark_entry(sessions, total_hours) if sessions else (None, None)

    return DayStatus(
        today_str=today_str,
        sessions=sessions,
        total_hours=total_hours,
        target_hours=TARGET_HOURS,
        active_start=active_start,
        active_hours=active_hours,
        active_task=active_task,
        promark_start=pm_start,
        promark_end=pm_end,
    )


def _close_session(time_override=None):
    """Set End on the open session. Returns (wb, ws) or (None, None)."""
    if not os.path.exists(EXCEL_FILE):
        print(_NO_FILE_MSG)
        return None, None

    wb = load_workbook(EXCEL_FILE)
    ws = wb[LOG_SHEET]

    open_row = find_open_session(ws)
    if not open_row:
        print("No open session found. Run 'start' first.")
        return None, None

    time_str = time_override or datetime.now().strftime("%H:%M")
    start_val = ws.cell(row=open_row, column=3).value
    date_val = ws.cell(row=open_row, column=1).value

    ws.cell(row=open_row, column=4, value=time_str)

    note = " (manually set)" if time_override else ""
    wb.save(EXCEL_FILE)
    print(f"⏸  Closed — {date_val}  {start_val} → {time_str}{note}")
    return wb, ws
