from datetime import date, datetime, timedelta
from openpyxl.utils import get_column_letter

from Constants import SUM_SHEET, TARGET_HOURS, C_HEADER, C_WHITE, C_OVER, C_UNDER, C_TODAY, C_ALT
from Styles import style_cell

PROMARK_SHEET = "Promark"
LUNCH_START = "11:45"
LUNCH_END = "12:15"
LUNCH_MINUTES = 30


def _parse_hhmm(s):
    return datetime.strptime(s, "%H:%M")


def _fmt_hhmm(dt):
    return dt.strftime("%H:%M")


def promark_entry(sessions, total_hours):
    """
    Given a list of (start, end) session tuples and the total logged hours,
    return (first_start, promark_end) as HH:MM strings.

    Promark end = first_start + total_hours + 30 min lunch.
    """
    if not sessions:
        return None, None

    first_start = _parse_hhmm(sessions[0][0])
    total_minutes = round(total_hours * 60)
    promark_end = first_start + timedelta(minutes=total_minutes + LUNCH_MINUTES)
    return _fmt_hhmm(first_start), _fmt_hhmm(promark_end)


def build_promark(wb):
    """Rebuild the Promark sheet from scratch."""
    from Workbook import read_log

    if PROMARK_SHEET in wb.sheetnames:
        del wb[PROMARK_SHEET]

    ws = wb.create_sheet(PROMARK_SHEET)
    from Constants import LOG_SHEET as _LOG
    ws_log = wb[_LOG]

    days = read_log(ws_log)
    if not days:
        ws["A1"] = "No completed sessions yet."
        return

    _write_promark_section(ws, days)
    ws.freeze_panes = "A2"


def _row_color(date_str):
    today_str = date.today().strftime("%Y-%m-%d")
    if date_str == today_str:
        return C_TODAY
    return C_ALT


def _write_promark_section(ws, days):
    headers = [
        "Date", "Week",
        "Enter in Promark: Start", "Enter in Promark: Lunch Start",
        "Enter in Promark: Lunch End", "Enter in Promark: End",
        "Actual Hours Logged", "Note"
    ]
    col_widths = [14, 8, 24, 24, 22, 24, 20, 40]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        style_cell(cell, bold=True, color=C_HEADER, font_color=C_WHITE, size=11)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22

    today_str = date.today().strftime("%Y-%m-%d")

    for i, (date_str, info) in enumerate(sorted(days.items()), start=2):
        sessions = info["sessions"]
        total = info["total"]
        week = info["week"]

        start_str, end_str = promark_entry(sessions, total)

        # Build a human-readable note
        session_parts = " + ".join(
            f"{s}→{e}" for s, e in sessions
        )
        n_sessions = len(sessions)
        if n_sessions == 1:
            note = f"Single session: {session_parts}"
        else:
            note = f"{n_sessions} sessions merged: {session_parts}"

        color = _row_color(date_str)

        values = [
            date_str, week,
            start_str, LUNCH_START, LUNCH_END, end_str,
            round(total, 2), note
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=col, value=val)
            num_fmt = "0.00" if col == 7 else None
            style_cell(cell, color=color, num_fmt=num_fmt)